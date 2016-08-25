# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from .models import Goods


def excel_reader(file):
    wb = load_workbook(file)
    ws = wb.active
    for row in ws.iter_rows():
        spisok = list(row)
        try:
            title_goods = spisok[0].value
            price_goods = spisok[1].value
            db_record = Goods.objects.get_or_create(title=title_goods, price=price_goods)
        except TypeError:
            print('поле пустое')